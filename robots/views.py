import json
import re
import datetime
import openpyxl

from calendar import monthrange

from django.shortcuts import redirect, reverse, get_object_or_404
from django.http.response import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.cache import cache
from django.core.mail import send_mail

from .models import RobotModelDB, RobotVersionDB, Robot, RobotOffer, RobotRelease
from customers.models import Customer
from orders.models import ActiveOrder, ReleaseOrder
# Create your views here.


regex = r'^[A-Z0-9]{2}$'

email_message= 'Добрый день!\nНедавно вы интересовались нашим роботом модели {}, версии {}.\nЭтот робот теперь в наличии. Если вам подходит этот вариант - пожалуйста, свяжитесь с нами\n'


@csrf_exempt
def robot_models_in_db(request):
    """
    Эта функция используется для внесения роботов в базу данных для
    их дальнейшего сравнения с поступающими запросами на создание
    роботов с целью валидации входных данных звпросов, а также для
    отображения всех моделей роботов,внесенных в базу данных предприятия.
    :param
    request: С помощью request.GET.get() происходит извлечение данных из ключей,
    распоагаемых в url, осуществляющих следующие действия:
    model - создает новую модель, например: "http://127.0.0.1:8000/models?model=R2";
    delete - удаляет модель из базы данных (крайне нежелательно),
    например: "http://127.0.0.1:8000/models?delete=R2".
    :return: Возвращает JSONResponse со списком моделей роботов, внесенных в базу данных
    """

    if request.GET.get("model"):
        model = request.GET.get("model")
        if not re.match(regex, model):
            return HttpResponse('model должна состоять из двух символов (Прописных '
                                'букв латинского алфавита и цифр, например: "R2", "RR", "22"')
        if not RobotModelDB.objects.filter(model=model):
            RobotModelDB.objects.create(model=model)
            return redirect(reverse('robots:models'))
        else:
            return HttpResponse('Для создание модели необходимо ввести уникальное значение, в '
                                'сдледующем формате: "http://127.0.0.1:8000/models?model=ЗНАЧЕНИЕ". '
                                'Например: "http://127.0.0.1:8000/models?model=R2"')

    if request.GET.get("delete"):
        model = request.GET.get("delete")
        if not re.match(regex, model):
            return HttpResponse('delete должна состоять из двух символов (Прописных '
                                'букв латинского алфавита и цифр, например: "R2", "RR", "22"')
        if not RobotModelDB.objects.filter(model=model):
            return HttpResponse('Данная версии не существует. Если хотите удалить версию '
                                'модели, введите в адресную строку '
                                '"http://127.0.0.1:8000/models?delete=ЗНАЧЕНИЕ". Крайне '
                                'нежелательное действие.')
        else:
            model = RobotModelDB.objects.get(model=model)
            model.delete()
            return redirect(reverse('robots:models'))

    data = list()
    model_list = RobotModelDB.objects.all()
    for model in model_list:
        model = dict(model=model.model,
                     hyperlink='http://127.0.0.1:8000/models/{}/'.format(model.pk))
        data.append(model)
    data = dict(total_models=model_list.count(),
                models=data)
    return JsonResponse(data, safe=False)


@csrf_exempt
def model_versions_in_db(request, pk):
    """
    Данное представление создано для того, что добавлять, обновлять или удалять
    версии определенной модели робота. Также данное представление отображает
    конкретную модель и версии модели для использования в качестве энпойнта.
    :param request: позволяет добавлять, обновлять и удалять версии модели путем введения
    в адресную строку: "http://127.0.0.1:8000/models/<id: модели>/" ключей:
    version - для создания новой версии модели, например: "http://127.0.0.1:8000/models/1?ver=T3"
    version, update - для обновления значения версии, например: "http://127.0.0.1:8000/models/1?ver=T3&upd=T5",
    delete - для удаления версии, например: "http://127.0.0.1:8000/models/1?del=T3".
    :param pk: Это id модели робота в базе данных предприятия
    :return: JSONResponse модели, общего числа версий модели и сами версии
    """
    model = get_object_or_404(RobotModelDB, pk=pk)

    if request.GET.get('version'):
        version = request.GET.get('version')
        if not re.match(regex, version):
            return HttpResponse('version должна состоять из двух символов (Прописных '
                                'букв латинского алфавита и цифр, например: "R2", "RR", "22"')
        if not model.versions.filter(version=version):
            model.versions.create(version=version)
            return redirect(reverse('robots:model_versions', kwargs={'pk': pk}))
        else:
            if request.GET.get('update'):
                upd = request.GET.get('update')
                if not re.match(regex, upd):
                    return HttpResponse('update должна состоять из двух символов (Прописных '
                                        'букв латинского алфавита и цифр, например: "R2", "RR", "22"')
                if not model.versions.filter(version=upd).exclude(version=version):
                    version_upd = RobotVersionDB.objects.get(model=model,
                                                             version=version)
                    version_upd.version = upd
                    version_upd.save()
                    return redirect(reverse('robots:model_versions', kwargs={'pk': pk}))
                else:
                    return HttpResponse("update должна содержать занчение версии, которой ещё "
                                        "не существует у данной модели")
            else:
                return HttpResponse('Данная версия уже существует. Если хотите изменить значение данной'
                                    'модели, введите в адресную строку '
                                    'http://127.0.0.1:8000/models/{}?version=ЗНАЧЕНИЕ&update=ЗНАЧЕНИЕ'.format(
                    model.pk
                ))

    if request.GET.get('delete'):
        version = request.GET.get('delete')
        if not re.match(regex, version):
            return HttpResponse('delete должна состоять из двух символов (Прописных '
                                'букв латинского алфавита и цифр, например: "R2", "RR", "22"')
        if model.versions.filter(version=version):
            version = RobotVersionDB.objects.get(model=model,
                                                 version=version)
            version.delete()
            return redirect(reverse('robots:model_versions', kwargs={'pk': pk}))
        else:
            return HttpResponse('Данная версии не существует. Если хотите удалить версию '
                                'модели, введите в адресную строку '
                                'http://127.0.0.1:8000/models/{}?delete=ЗНАЧЕНИЕ'.format(
                model.pk
            ))

    versions = list()
    for version in model.versions.all():
        version = dict(version=version.version,
                       hyperlink='http://127.0.0.1:8000/models/{}/versions/{}/'.format(
                           model.pk, version.pk
                       ))
        versions.append(version)
    model = dict(model=model.model,
                 total_version=model.versions.all().count(),
                 versions=versions)

    return JsonResponse(model, safe=False)


@csrf_exempt
def versions_in_db(request, pk):
    """
    Данная функция используется для отображения версий той или иной модели робота,
    а также для добавления, обновления и удаления версий модели посредством запроса
    на url.
    :param request: позволяет добавлять, обновлять и удалять версии модели путем введения
    в адресную строку: "http://127.0.0.1:8000/models/<id: модели>/versions" ключей:
    ver - для создания новой версии модели, например: "http://127.0.0.1:8000/models/versions?ver=T3"
    ver, upd - для обновления значения версии, например: "http://127.0.0.1:8000/models/versions?ver=T3&upd=T5",
    del - для удаления версии, например: "http://127.0.0.1:8000/models/versions?del=T3".
    :param pk: это id модели робота, версии которого будут отображаться в JSON-формате
    :return: JSONResponse, который будет отражать версии модели в формате JSON.
    """
    model = get_object_or_404(RobotModelDB, pk=pk)

    if request.GET.get('version'):
        version = request.GET.get('version')
        if not re.match(regex, version):
            return HttpResponse('version должна состоять из двух символов (Прописных '
                                'букв латинского алфавита и цифр, например: "R2", "RR", "22"')
        if not model.versions.filter(version=version):
            model.versions.create(version=version)
            return redirect(reverse('robots:versions', kwargs={'pk': model.pk}))
        else:
            if request.GET.get('update'):
                upd = request.GET.get('update')
                if not re.match(regex, upd):
                    return HttpResponse('update должна состоять из двух символов (Прописных '
                                        'букв латинского алфавита и цифр, например: "R2", "RR", "22"')
                if not model.versions.filter(version=upd).exclude(version=version):
                    version = RobotVersionDB.objects.get(model=model,
                                                         version=version)
                    version.version = upd
                    version.save()
                    return redirect(reverse('robots:versions', kwargs={'pk': model.pk}))
                else:
                    return HttpResponse("update должна содержать занчение версии, которой ещё "
                                        "не существует у данной модели")
            else:
                return HttpResponse('Данная версия уже существует. Если хотите изменить значение данной'
                                    'модели, введите в адресную строку '
                                    'http://127.0.0.1:8000/models/{}/versions?version=ЗНАЧЕНИЕ&update=ЗНАЧЕНИЕ'.format(
                    model.pk
                ))
    if request.GET.get('delete'):
        version = request.GET.get('delete')
        if not re.match(regex, version):
            return HttpResponse('delete должна состоять из двух символов (Прописных '
                                'букв латинского алфавита и цифр, например: "R2", "RR", "22"')
        if model.versions.filter(version=version):
            version = model.versions.get(version=version)
            version.delete()
            return redirect(reverse('robots:versions', kwargs={'pk': model.pk}))
        else:
            return HttpResponse('Данная версии не существует. Если хотите удалить версию '
                                'модели, введите в адресную строку '
                                'http://127.0.0.1:8000/models/{}/versions?delete=ЗНАЧЕНИЕ'.format(
                model.pk
            ))

    data = list()
    for version in model.versions.all():
        version = dict(version=version.version,
                       hyperlink=f'http://127.0.0.1:8000/models/{model.pk}/versions/{version.pk}/')
        data.append(version)
    data = dict(versions=data)
    return JsonResponse(data, safe=False)


@csrf_exempt
def serial_number_in_db(request, id, pk):
    """
    Данная функция отображает конкретный серийный номмер конкретной модели
    робота, занесенного в базу данных предприятия. На данный эндпойнт можно
    ссылаться для проверки валидности того или иного серийного номера.
    :param request: позволяет удалить данный серийный номер из базы данных,
    добавив "delete" в url формата:"http://127.0.0.1:8000/models/<id: модели>/version/<id: версии>/",
    например: "http://127.0.0.1:8000/models/3/version/5?delete=yes"
    :param id:это id версии робота
    :param pk: это id модели робота
    :return: JSONResponse, отображающий серийный номер робота, в формате JSON.
    """
    model = get_object_or_404(RobotModelDB, pk=pk)
    version = get_object_or_404(RobotVersionDB, pk=id)

    if request.GET.get("delete"):
        if request.GET.get("delete") == 'yes':
            version.delete()
            return redirect(reverse('robots:model_versions', kwargs={'pk': model.pk}))
        else:
            return HttpResponse('Чтобы удалить данную версию модели, введите в адресную строку '
                                '"http://127.0.0.1:8000/models/{}/version/{}?delete=yes"'.format(
                model.pk,
                version.pk
            ))

    data = dict(model=model.model,
                version=version.version)
    return JsonResponse(data)


def total_robots_serial_in_db(request):
    """
    Эта функция используется для подробного отражения всех моделей и их версий,
    внесенных в базу данных предприятия в на одной странице.
    :param request:
    :return:
    Сама view отображает все уникалькальные модели роботов, которые
    содержатся в базе данных предприятия вместе с их версиями, серийными номерами
    и ссылками на каждую модель и версию.
    """

    data = list()
    model_list = RobotModelDB.objects.all()
    for model in model_list:
        versions = list()
        for version in model.versions.all():
            version = dict(version=version.version,
                           serial='{}-{}'.format(model.model, version.version),
                           hyperlink='http://127.0.0.1:8000/models/{}/versions/{}/'.format(
                               model.pk, version.pk
                           ))
            versions.append(version)
        model = dict(model=model.model,
                     hyperlink='http://127.0.0.1:8000/models/{}/'.format(model.pk),
                     versions=versions)
        data.append(model)

    data = dict(total_models=model_list.count(),
                total_versions=RobotVersionDB.objects.all().count(),
                заказчики='http://127.0.0.1:8000/customers/',
                все_ордеры='http://127.0.0.1:8000/orders/',
                активные_ордеры='http://127.0.0.1:8000/orders/actives/',
                реализованные_ордера='http://127.0.0.1:8000/orders/releases/',
                производство_роботов='http://127.0.0.1:8000/robots/',
                предложения_роботов='http://127.0.0.1:8000/robots/offers/',
                проданные_роботы='http://127.0.0.1:8000/robots/releases/',
                отчет_за_неделю='http://127.0.0.1:8000/robots/week/',
                models=data)
    return JsonResponse(data, safe=False)


@csrf_exempt
def robots_production(request):
    """
    Данная функция добавляет данные о произведенных роботах, посредвом
    специального обращения к данному эндпойнту, путем добавления серийного номера
    в адрессную строку браузера к url. Также, при создании робота, если он есть в активных
    заказах, ордер заказчика переводится из активного в неактивный, а самому заказчику на email
    отправляется сообщение об изготовлении робота.
    :param request: обеспечивает извлечение данных, переданных в параметре serail посредством
    request.GET.get, после чего проводится валидация данных на предмет соответствия серийным номерам,
    представленных в базе данных предприятия.
    Значение, передаваемое в сериал, должно соответствовать формату "два символа-два символа".
    Допускается использование только прописных букв латинского алфавита и цифры.
    Например: http://127.0.0.1:8000/robots?serial=R2-D2". В случае некоректно введенных данных,
    происходит перенаправление на эндпойнт "http://127.0.0.1:8000", где представлены все
    модели роботов с их версиями, зарегистрированные в базе данных предприятия.
    :return:JSONResponse со словарем, содержащим список всех произведенных роботов, со временем их
    изготовления.

    P.S. Данная функция может работать некорректно в браузере Chrome, поэтому, переда работой с
    данной функцией необходимо удалить историю браузера за последнее время.
    """

    from .scripts import to_order_robot_production

    regex_serial = r'^[A-Z0-9]{2}[\-]{1}[A-Z0-9]{2}$'
    regex_datetime = r'^[0-9]{4}[\s]{1}[0-9]{2}[\s]{1}[0-9]{2}[\s]{1}[0-9]{2}[\s]{1}[0-9]{2}[\s]{1}[0-9]{2}$'

    cache.delete('http://127.0.0.1:8000/robots/')
    cache.delete('http://127.0.0.1:8000/robots?serial=')

    if request.GET.get('serial'):
        serial = request.GET.get('serial')
        if not re.match(regex_serial, serial):
            return HttpResponse('Введите серийный номер в формате: '
                                '2 символа - дефис- 2 символа '
                                '(допускаются прописные буквы латинского алфавита и цифры. Например:'
                                '"T3-L0". '
                                'Серийные номера можно посмотреть по адресу: http://127.0.0.1:8000/')
        serial = serial.split('-')
        model = serial[0]
        version = serial[1]
        if not RobotModelDB.objects.filter(model=model):
            return HttpResponse('Данная модель отсутствует в каталоге предприятия. Со всеми'
                                'моделями и серийными номерами можно ознакомиться по адресу: '
                                'http://127.0.0.1:8000/models/')
        else:
            robot = RobotModelDB.objects.get(model=model)
            if not robot.versions.filter(version=version):
                return HttpResponse('Данная версия отсутсвует у выбранной модели робота.'
                                    'Узнать все версии модели можно по адресу: '
                                    'http://127.0.0.1:8000/models/{}/'.format(robot.pk))
            else:
                if request.GET.get('created'):
                    created = request.GET.get('created')
                    if not re.match(regex_datetime, created):
                        return HttpResponse('Дата создания должна быть написана в следующем формате, '
                                            'например: '
                                            '"http://127.0.0.1:8000/robots?serial=R2-D2&created=2024+03+01+13+34+45"')
                    created = created.split()
                    year = int(created[0])
                    month = int(created[1])
                    day = int(created[2])
                    hour = int(created[3])
                    minute = int(created[4])
                    seconds = int(created[5])
                    if year > datetime.datetime.now().year:
                        return HttpResponse('Год производства не может быть в будущем')
                    elif 0 > month > 12:
                        return HttpResponse('Месяц должен быть представлен в формате от "01" до "12')
                    elif 0 > day > monthrange(year, month)[1]:
                        return HttpResponse('Укажите число месяца')
                    elif 0 > hour > 24:
                        return HttpResponse('Укажите время в районе 24 часов')
                    elif 0 > minute > 60:
                        return HttpResponse('Укажите время в районе 60 минут')
                    elif 0 > seconds > 60:
                        return HttpResponse('Укажите время в районе 60 секунд')
                    else:
                        created = f'{year}-{month}-{day} {hour}:{minute}:{seconds}'
                        Robot.objects.create(model=model,
                                             version=version,
                                             created=datetime.datetime.strptime(created, '%Y-%m-%d %H:%M:%S'))
                        robot = RobotOffer.objects.create(model=model,
                                                          version=version,
                                                          created=datetime.datetime.strptime(created, '%Y-%m-%d %H:%M:%S'))

                        if not ActiveOrder.objects.filter(robot_serial=robot.serial):
                            return redirect(reverse('robots:production'))
                        else:
                            order = ActiveOrder.objects.filter(robot_serial=robot.serial).first()
                            customer = Customer.objects.get(pk=order.customer_id)
                            RobotRelease.objects.create(model=robot.model,
                                                        version=robot.version,
                                                        created=robot.created)
                            ReleaseOrder.objects.create(customer=customer,
                                                        robot_serial=order.robot_serial)

                            send_mail(
                                subject='Ваш заказ готов',
                                message=email_message.format(robot.model, robot.version),
                                from_email='vitalytanceforwork@gmail.com',
                                recipient_list=[customer.email]
                            )
                            robot.delete()
                            order.delete()
                            return redirect(reverse('robots:production'))

                else:
                    Robot.objects.create(model=model,
                                         version=version,
                                         created=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                    robot = RobotOffer.objects.create(model=model,
                                                      version=version,
                                                      created=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

                    if not ActiveOrder.objects.filter(robot_serial=robot.serial):
                        return redirect(reverse('robots:production'))
                    else:
                        order = ActiveOrder.objects.filter(robot_serial=robot.serial).first()
                        customer = Customer.objects.get(pk=order.customer_id)
                        RobotRelease.objects.create(model=robot.model,
                                                    version=robot.version,
                                                    created=robot.created)
                        ReleaseOrder.objects.create(customer=customer,
                                                    robot_serial=order.robot_serial)

                        send_mail(
                            subject='Ваш заказ готов',
                            message=email_message.format(robot.model, robot.version),
                            from_email='vitalytanceforwork@gmail.com',
                            recipient_list=[customer.email]
                        )
                        robot.delete()
                        order.delete()
                        return redirect(reverse('robots:production'))

    data = list()
    robots = Robot.objects.all()
    for robot in robots:
        robot = dict(model=robot.model,
                     version=robot.version,
                     created=robot.created.strftime('%Y-%m-%d %H:%M:%S'))
        data.append(robot)
    data.sort(key=to_order_robot_production)
    data = dict(production=data)

    actives = ActiveOrder.objects.all()

    for active in actives:
        if RobotOffer.objects.filter(serial=active.robot_serial):
            robot = RobotOffer.objects.filter(serial=active.robot_serial).first()
            RobotRelease.objects.create(model=robot.model,
                                        version=robot.version,
                                        created=robot.created)
            customer = Customer.objects.get(pk=active.customer_id)
            ReleaseOrder.objects.create(customer=active.customer,
                                        robot_serial=active.robot_serial)
            send_mail(
                subject='Ваш заказ готов',
                message=email_message.format(robot.model, robot.version),
                from_email='vitalytanceforwork@gmail.com',
                recipient_list=[customer.email]
            )
            robot.delete()
            active.delete()

    return JsonResponse(data, safe=False)


@csrf_exempt
def robots_offers(request):
    """
    Данное представление отображает все нереализованные произведенные роботы.
    При создании робота, робот располагается в двух таблицах - "robots_robot" и "robots_robotoffer".
    Сразу же после производства и добавления в в список предложений, происходит проверка относительно
    ордеров от потребителей. И если находится соответсвие с заявкой, то робот добавляется в таблицу
    "robots_robotrelease", а из robotoffer адаляется соответсвующая запись. Сам ордер перемешается из
    таблицы активных заявок в реализованные, а пользователю отправляется сообщение о произведенном роботе.
    :param request: Обращение к представлению
    :return:JSONResponse c предложениями роботов.
    """

    offers = RobotOffer.objects.all().values('model', 'version', 'created', 'serial')
    data = list()

    for offer in offers:
        data.append(offer)

    data = dict(offers=data)

    return JsonResponse(data, safe=False)


@csrf_exempt
def robots_releases(request):
    """
    Данное представление отображает все реализованные произведенные роботы.
    При создании робота, робот располагается в двух таблицах - "robots_robot" и "robots_robotoffer".
    Сразу же после производства и добавления в в список предложений, происходит проверка относительно
    ордеров от потребителей. И если находится соответсвие с заявкой, то робот добавляется в таблицу
    "robots_robotrelease", а из robotoffer адаляется соответсвующая запись. Сам ордер перемешается из
    таблицы активных заявок в реализованные, а пользователю отправляется сообщение о произведенном роботе.
    :param request: Обращение к представлению
    :return:JSONResponse cо списком json-записей реализованных роботов.
    """

    releases = RobotRelease.objects.all().values('model', 'version', 'created', 'serial')
    data = list()

    for release in releases:
        data.append(release)

    data = dict(releases=data)

    return JsonResponse(data, safe=False)


def week_report(request):
    """
    Данная функция создает отчет о всех сделанных моделях работов и их версиях
    за прошедшиую неделю в виде xlsx документа, который содержит страницы по каждой модели,
    находящейся в базе данных предприятия.
    К сожалению, я не нашел способа сделать данный документ стандартными средствами Django
    и Python, поэтому пришлось использовать библиотеку openpyxl
    :param request: в данной фунции не используется. Необходим лишь для вызова функции по
    обращению к данному эндпойнту.
    :return: response, содержащий в себе xlsx документ.
    """
    wb = openpyxl.Workbook()

    week_ago = datetime.datetime.today() - datetime.timedelta(days=7)
    models = RobotModelDB.objects.all()

    for model in models:
        ws = wb.create_sheet(f'{model.model}')
        versions = model.versions.all()
        data = list()

        for version in versions:
            robots = Robot.objects.filter(model=model.model,
                                          version=version.version,
                                          created__range=[week_ago, datetime.datetime.today()]).count()
            robots = dict(A=model.model,
                          B=version.version,
                          C=robots)
            data.append(robots)

        ws['A1'] = 'Модель'
        ws['B1'] = 'Версия'
        ws['C1'] = 'Количество за неделю'

        for row in data:
            ws.append(row)

    week_ago = week_ago.strftime('%Y-%m-%d')
    today = datetime.datetime.today().strftime('%Y-%m-%d')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=week report ({week_ago})-({today}).xlsx'

    wb.remove(wb['Sheet'])
    wb.save(response)

    return response
